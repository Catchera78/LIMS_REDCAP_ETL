"""Lecture d'un onglet Excel LIMS en une grille (list[list[str]]).

Strategie a deux niveaux :
  1. openpyxl si disponible (recommande, voir requirements.txt) ;
  2. sinon, repli 100 % bibliotheque standard (zipfile + XML) afin que le
     pipeline reste executable sur un poste ou openpyxl n'est pas encore
     installe.

Le fichier original n'est JAMAIS modifie (ouverture en lecture seule).
La transformation des donnees n'est PAS faite ici (Prompt 1) : on ne fait
que lire la structure et le contenu brut sous forme de chaines.
"""
from __future__ import annotations

import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import List, Optional, Tuple

from .text_utils import cell_to_str, normalize

try:  # openpyxl est optionnel a l'execution (present en prod via requirements.txt)
    import openpyxl  # type: ignore
    _HAVE_OPENPYXL = True
except Exception:  # pragma: no cover - depend de l'environnement
    _HAVE_OPENPYXL = False


class ExcelReadError(Exception):
    """Erreur de lecture du fichier Excel LIMS."""


# --------------------------------------------------------------------------- #
# Detection du fichier d'entree
# --------------------------------------------------------------------------- #
def find_input_xlsx(input_dir: Path) -> Path:
    """Retourne l'unique fichier .xlsx present dans input/.

    - 0 fichier   -> ExcelReadError ;
    - 1 fichier   -> ok ;
    - >1 fichier  -> ExcelReadError (l'utilisateur doit n'en laisser qu'un).

    Les fichiers temporaires Excel (~$...) sont ignores.
    """
    input_dir = Path(input_dir)
    if not input_dir.is_dir():
        raise ExcelReadError(f"Dossier d'entree introuvable : {input_dir}")

    candidates = [
        p for p in sorted(input_dir.glob("*.xlsx"))
        if not p.name.startswith("~$")
    ]
    if not candidates:
        raise ExcelReadError(
            f"Aucun fichier .xlsx trouve dans {input_dir}. "
            "Placez l'extraction LIMS dans ce dossier."
        )
    if len(candidates) > 1:
        noms = ", ".join(p.name for p in candidates)
        raise ExcelReadError(
            f"Plusieurs fichiers .xlsx dans {input_dir} ({noms}). "
            "Ne laissez qu'une seule extraction a la fois."
        )
    return candidates[0]


# --------------------------------------------------------------------------- #
# Selection de l'onglet
# --------------------------------------------------------------------------- #
def _match_sheet(names: List[str], sheet_name: str,
                 fallback_contains: Optional[str]) -> str:
    """Choisit le nom d'onglet : exact, sinon normalise, sinon 'contient'."""
    # 1) correspondance exacte
    if sheet_name in names:
        return sheet_name
    # 2) correspondance normalisee (accents/casse tolerants)
    target = normalize(sheet_name)
    for n in names:
        if normalize(n) == target:
            return n
    # 3) repli : onglet dont le nom normalise contient le motif
    if fallback_contains:
        needle = normalize(fallback_contains)
        for n in names:
            if needle in normalize(n):
                return n
    raise ExcelReadError(
        f"Onglet '{sheet_name}' introuvable. Onglets disponibles : {names}"
    )


# --------------------------------------------------------------------------- #
# Lecture -> grille
# --------------------------------------------------------------------------- #
def read_grid(path: Path, sheet_name: str,
              fallback_contains: Optional[str] = None) -> Tuple[List[List[str]], str, str]:
    """Lit l'onglet demande et renvoie (grille, nom_onglet_reel, moteur).

    - grille : list de lignes, chaque ligne = list de chaines (cellules vides
      = "") ; les lignes sont completees a la largeur maximale.
    - moteur : "openpyxl" ou "stdlib" (trace dans le log).
    """
    path = Path(path)
    if not path.is_file():
        raise ExcelReadError(f"Fichier Excel introuvable : {path}")

    if _HAVE_OPENPYXL:
        grid, real = _read_grid_openpyxl(path, sheet_name, fallback_contains)
        return grid, real, "openpyxl"
    grid, real = _read_grid_stdlib(path, sheet_name, fallback_contains)
    return grid, real, "stdlib"


def _pad(grid: List[List[str]]) -> List[List[str]]:
    width = max((len(r) for r in grid), default=0)
    for r in grid:
        if len(r) < width:
            r.extend([""] * (width - len(r)))
    return grid


def _read_grid_openpyxl(path: Path, sheet_name: str,
                        fallback_contains: Optional[str]):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        real = _match_sheet(list(wb.sheetnames), sheet_name, fallback_contains)
        ws = wb[real]
        grid: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            grid.append([cell_to_str(v) for v in row])
        return _pad(grid), real
    finally:
        wb.close()


# ------- Repli bibliotheque standard : parse minimal du format XLSX -------- #
_MAIN_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"


def _col_index(cell_ref: str) -> int:
    """A1 -> 0, B1 -> 1, AA1 -> 26 ... (0-based)."""
    letters = re.match(r"([A-Z]+)", cell_ref).group(1)
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


def _normalize_part(target: str) -> str:
    """Normalise une cible de relation workbook -> chemin dans le zip.

    - cible absolue "/xl/worksheets/sheet2.xml" -> "xl/worksheets/sheet2.xml" ;
    - cible relative "worksheets/sheet2.xml"     -> "xl/worksheets/sheet2.xml" ;
    - cible deja prefixee "xl/worksheets/..."     -> inchangee.
    """
    if target.startswith("/"):
        return target.lstrip("/")           # absolue depuis la racine du paquet
    if target.startswith("xl/"):
        return target
    return "xl/" + target                    # relative au dossier xl/


def _read_grid_stdlib(path: Path, sheet_name: str,
                      fallback_contains: Optional[str]):
    with zipfile.ZipFile(path) as z:
        # shared strings
        shared: List[str] = []
        if "xl/sharedStrings.xml" in z.namelist():
            sroot = ET.fromstring(z.read("xl/sharedStrings.xml"))
            for si in sroot.iter(_MAIN_NS + "si"):
                shared.append("".join(t.text or "" for t in si.iter(_MAIN_NS + "t")))

        # noms d'onglets -> rId
        wbroot = ET.fromstring(z.read("xl/workbook.xml"))
        name_to_rid = {}
        for s in wbroot.iter(_MAIN_NS + "sheet"):
            name_to_rid[s.get("name")] = s.get(_REL_NS + "id")

        real = _match_sheet(list(name_to_rid), sheet_name, fallback_contains)

        # rId -> cible (fichier worksheet). La cible peut etre :
        #   - relative au dossier xl/ : "worksheets/sheet2.xml" ;
        #   - absolue depuis la racine : "/xl/worksheets/sheet2.xml".
        relroot = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {r.get("Id"): r.get("Target") for r in relroot}
        target = _normalize_part(rid_to_target[name_to_rid[real]])

        sroot = ET.fromstring(z.read(target))
        grid: List[List[str]] = []
        for row in sroot.iter(_MAIN_NS + "row"):
            cells = {}
            for c in row.iter(_MAIN_NS + "c"):
                ref = c.get("r")
                idx = _col_index(ref) if ref else len(cells)
                ctype = c.get("t")
                v = c.find(_MAIN_NS + "v")
                is_ = c.find(_MAIN_NS + "is")
                if ctype == "s" and v is not None:
                    val = shared[int(v.text)]
                elif is_ is not None:
                    val = "".join(t.text or "" for t in is_.iter(_MAIN_NS + "t"))
                elif v is not None:
                    val = v.text or ""
                else:
                    val = ""
                cells[idx] = val
            if cells:
                width = max(cells) + 1
                grid.append([cells.get(i, "") for i in range(width)])
            else:
                grid.append([])
        return _pad(grid), real


def list_sheet_names(path: Path) -> List[str]:
    """Liste les onglets (utile pour diagnostic)."""
    path = Path(path)
    if _HAVE_OPENPYXL:
        wb = openpyxl.load_workbook(path, read_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    with zipfile.ZipFile(path) as z:
        wbroot = ET.fromstring(z.read("xl/workbook.xml"))
        return [s.get("name") for s in wbroot.iter(_MAIN_NS + "sheet")]


def engine_name() -> str:
    return "openpyxl" if _HAVE_OPENPYXL else "stdlib"
